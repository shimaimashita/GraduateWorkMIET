import os
from datetime import datetime
from pathlib import Path

import openpyxl
import smtplib

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

import mimetypes


from django.conf import settings
from django.conf.urls.static import static
from django.contrib import messages
from django.shortcuts import render, redirect
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic.edit import CreateView
from django.http.response import HttpResponse

from .models import Faculty, Group, Mailing, Question, Result, Student
from .models import Upload
from .forms import QuestionForm, MailingForm


def post_new(request):
    if request.method == "POST":
        form = QuestionForm(request.POST)
        if form.is_valid():
            post = form.save(commit=False)
            post.year = datetime.today()
            post.save()
            return redirect('/')
    else:
        form = QuestionForm()
    return render(request, 'asking/question_edit.html', {'form': form})

def statistic_new(request):
    file = 'result.xlsx'
    filepath = Path(settings.MEDIA_ROOT, file)
    if request.method == "POST":
        wb = Workbook()
        ws = wb.active
        ws.append(['Всего', 'Трудоустроенных', 'Нетрудоустроенных c по причине', 'Остальные'])
        data = list()
        all_students = Question.objects.all().count()
        employed = Question.objects.filter(employed__exact='y').count()
        unemployed = Question.objects.filter(studying__exact='y').exclude(employed__exact='y').count()
        other = all_students - employed - unemployed
        data = [all_students, employed, unemployed, other]
        ws.append(data)
        wb.save(filepath)

        with open(filepath, 'rb') as fh:
            response = HttpResponse(fh.read(),
                                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = 'attachment; filename=result.xlsx'
            return response
    return render(request, 'asking/statistic_new.html')


def mailing_new(request):
    if request.method == "POST":
        form = MailingForm(request.POST)
        if form.is_valid():
            post = form.save(commit=False)
            num_of = Student.objects.filter(graduate_year__year=post.get_year()).count()
            context = {
                'num_of': num_of
            }
            if num_of > 0:
                email = 'shimai.mashita@yandex.ru'
                password = 'xjkqnnnzprbxporo'

                server = smtplib.SMTP('smtp.yandex.ru', 587)
                server.ehlo()
                server.starttls()
                server.login(email, password)

                dest_email = 'kchuslyaev@mail.ru'
                subject = 'MIET'
                email_text = 'http://127.0.0.1:8000/asking/asking/new/'
                message = 'From: %s\nTo: %s\nSubject: %s\n\n%s' % (email, dest_email, subject, email_text)

                server.set_debuglevel(1)
                server.sendmail(email, dest_email, message)
                server.quit()

            return render(request, 'asking/mailing.html', context=context)
    else:
        form = MailingForm()
    return render(request, 'asking/mailing.html', {'form': form})

def index(request):
    """View function for home page of site."""

    # Generate counts of some of the main objects
    num_of_answers = 10
    num_of_completed_mailing = 1
    all_faculties = ' ,'.join(faculty.faculty_name for faculty in Faculty.objects.all())
    all_groups = ' ,'.join(group.group_name for group in Group.objects.all())
    all_question = Question.objects.all().count()
    employed = Question.objects.filter(employed__exact='y').count()

    context = {
        'num_of_answers': num_of_answers,
        'num_of_completed_mailing': num_of_completed_mailing,
        'all_faculties': all_faculties,
        'all_groups': all_groups,
        'all_question': all_question,
        'employed': employed
    }

    # Render the HTML template index.html with the data in the context variable
    return render(request, 'index.html', context=context)


class UploadView(CreateView):
    model = Upload
    fields = ['upload_file', ]
    success_url = reverse_lazy('fileupload')

    def get_context_data(self, **kwargs):

        context = super().get_context_data(**kwargs)
        context['documents'] = Upload.objects.all().filter(file_checked__exact=False)
        result_str = list()
        for document in context['documents']:
            xslx_file = document.upload_file.name
            wb_obj = openpyxl.load_workbook(Path(settings.MEDIA_ROOT, xslx_file))
            sheet = wb_obj.active
            for row in sheet.iter_rows(min_row=2):
                temp_str = ''
                mail = row[1].value
                group = row[2].value

                if isinstance(mail, str) and isinstance(group, str):
                    try:
                        obj_group = Group.objects.get(group_name=group)
                    except Group.DoesNotExist:
                        obj_group = Group(group_name=group)
                    # except Student.DoesNotExist:
                    try:
                        obj = Student.objects.get(mail=mail)
                    except Student.DoesNotExist:
                        obj = Student(mail=mail, group=obj_group)
                    document.file_checked = True
                    document.save()
                    obj.save()
        context['num_of_completed'] = Upload.objects.all().filter(file_checked__exact=False).count()
        return context
